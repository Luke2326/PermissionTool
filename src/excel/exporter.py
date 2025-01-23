import logging
from datetime import datetime
import psycopg2
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from pathlib import Path
from typing import Dict, List, Generator, Optional
import concurrent.futures
from tqdm import tqdm
import io
import threading
from queue import Queue
import time
import sys
import os
import gc  # Importa il modulo gc per gestione memoria

# Configura il logging per essere sempre visibile
class TqdmToLogger(io.StringIO):
    logger = None
    level = None
    buf = ''
    def __init__(self, logger, level=None):
        super(TqdmToLogger, self).__init__()
        self.logger = logger
        self.level = level or logging.INFO
    
    def write(self, buf):
        self.buf = buf.strip('\r\n\t ')
        if self.buf:
            self.logger.log(self.level, self.buf)
    
    def flush(self):
        if self.buf:
            self.logger.log(self.level, self.buf)
        self.buf = ''

# Configura il logger per scrivere su stdout con flush immediato
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger()
tqdm_logger = TqdmToLogger(logger, level=logging.INFO)

def log_info(message: str):
    """Wrapper per logging che forza il flush"""
    logger.info(message)
    sys.stdout.flush()

# Tabelle che verranno gestite separatamente per performance
LARGE_TABLES = {'permissions', 'set_role_group_version'}

class ExcelWriter:
    def __init__(self):
        self.header_style = {
            'fill': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        self.main_file = None
        self.temp_dir = None
        self.large_files = {}

    def initialize(self, output_path: str):
        """Inizializza l'ambiente di scrittura"""
        self.main_file = output_path
        self.temp_dir = Path(output_path).parent / "temp_excel"
        self.temp_dir.mkdir(exist_ok=True)
        
        log_info(f"Inizializzazione ambiente di scrittura in {output_path}")
        wb = Workbook()
        wb.save(self.main_file)
        log_info("File Excel principale creato")

    def _apply_header_style(self, sheet):
        """Applica stile all'header"""
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = self.header_style['fill']
            cell.font = self.header_style['font']
            cell.border = self.header_style['border']
            cell.alignment = self.header_style['alignment']

    def _optimize_column_widths(self, sheet, sample_size=100):
        """Ottimizza larghezza colonne"""
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = chr(64 + col)
            
            # Campiona solo alcune righe
            for row in range(1, min(sample_size, sheet.max_row + 1)):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            sheet.column_dimensions[column].width = min(max_length + 2, 50)

    def write_dataframe(self, df: pd.DataFrame, sheet_name: str):
        """Scrive il DataFrame nel file appropriato"""
        start_time = time.time()
        total_rows = len(df)
        
        try:
            if sheet_name.lower() in LARGE_TABLES:
                # Per tabelle grandi (es. Permissions con ~380k righe)
                csv_file = self.temp_dir / f"{sheet_name}.csv"
                log_info(f"Inizio scrittura {sheet_name} ({total_rows} righe) in CSV temporaneo...")
                
                # Scrivi direttamente in CSV
                df.to_csv(csv_file, index=False)
                
                self.large_files[sheet_name] = csv_file
                log_info(f"CSV {sheet_name} completato")
                
                # Libera memoria
                del df
                gc.collect()
                
            else:
                log_info(f"Inizio scrittura {sheet_name} ({total_rows} righe) nel file principale...")
                wb = load_workbook(self.main_file)
                
                if sheet_name in wb.sheetnames:
                    wb.remove(wb[sheet_name])
                
                sheet = wb.create_sheet(sheet_name)
                
                # Scrivi header
                for col, name in enumerate(df.columns, 1):
                    sheet.cell(row=1, column=col, value=str(name))
                
                # Scrivi dati in batch
                batch_size = 1000
                for i in range(0, len(df), batch_size):
                    batch = df.iloc[i:i + batch_size]
                    for row_idx, row in enumerate(batch.values, i + 2):
                        for col_idx, value in enumerate(row, 1):
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    if (i + batch_size) % 10000 == 0:
                        log_info(f"Scritte {i + batch_size} righe di {total_rows} per {sheet_name}")
                        gc.collect()
                
                self._apply_header_style(sheet)
                self._optimize_column_widths(sheet)
                
                wb.save(self.main_file)
                wb.close()
                
                # Libera memoria
                del df
                gc.collect()
            
            elapsed = time.time() - start_time
            log_info(f"Elaborazione {sheet_name} completata in {elapsed:.2f} secondi")
            
        except Exception as e:
            log_info(f"Errore durante la scrittura di {sheet_name}: {str(e)}")
            raise

    def finalize(self):
        """Unisce tutti i file in uno solo"""
        if not self.large_files:
            return

        log_info("Inizio fase di unione dei file...")
        start_time = time.time()

        try:
            wb = load_workbook(self.main_file)
            
            for sheet_name, csv_file in self.large_files.items():
                try:
                    log_info(f"Processando {sheet_name}...")
                    
                    # Rimuovi il foglio se esiste
                    if sheet_name in wb.sheetnames:
                        wb.remove(wb[sheet_name])
                    
                    # Crea un nuovo foglio
                    sheet = wb.create_sheet(sheet_name)
                    
                    # Leggi il CSV in chunks
                    chunk_size = 5000
                    current_row = 1
                    header_written = False
                    
                    with open(csv_file, 'r') as f:
                        total_rows = sum(1 for _ in f) - 1  # -1 per header
                    
                    log_info(f"Iniziando importazione di {total_rows} righe per {sheet_name}")
                    
                    for chunk in pd.read_csv(csv_file, chunksize=chunk_size):
                        if not header_written:
                            # Scrivi header
                            for col, name in enumerate(chunk.columns, 1):
                                sheet.cell(row=1, column=col, value=str(name))
                            current_row = 2
                            header_written = True
                        
                        # Scrivi dati
                        for _, row in chunk.iterrows():
                            for col_idx, value in enumerate(row, 1):
                                try:
                                    sheet.cell(row=current_row, column=col_idx, value=value)
                                except Exception as cell_error:
                                    log_info(f"Errore nella scrittura della cella [{current_row}, {col_idx}]: {str(cell_error)}")
                                    raise
                            current_row += 1
                        
                        # Salva e log periodico
                        if current_row % 50000 == 0:
                            log_info(f"Processate {current_row-2} righe di {total_rows} per {sheet_name}")
                            wb.save(self.main_file)
                            gc.collect()
                    
                    # Applica stili
                    self._apply_header_style(sheet)
                    self._optimize_column_widths(sheet)
                    
                    # Salva
                    log_info(f"Salvataggio foglio {sheet_name}")
                    wb.save(self.main_file)
                    
                    # Rimuovi CSV
                    os.remove(csv_file)
                    log_info(f"File CSV {sheet_name} rimosso")
                    
                except Exception as e:
                    log_info(f"Errore durante l'elaborazione di {sheet_name}: {str(e)}")
                    raise
            
            # Chiudi workbook
            wb.close()
            
            # Pulizia directory temporanea
            try:
                os.rmdir(self.temp_dir)
                log_info("Directory temporanea rimossa")
            except Exception as e:
                log_info(f"Errore durante la rimozione della directory temporanea: {str(e)}")
            
            elapsed = time.time() - start_time
            log_info(f"Unione completata in {elapsed:.2f} secondi")
            
        except Exception as e:
            log_info(f"Errore durante la fase di unione: {str(e)}")
            raise

class DatabaseFetcher:
    CHUNK_SIZE = 100000

    def __init__(self, config: Dict):
        self.config = config
        self.conn = None

    def connect(self) -> bool:
        try:
            log_info(f"Connessione al database {self.config.get('database')} su {self.config.get('host')}...")
            self.conn = psycopg2.connect(**self.config)
            self.conn.set_session(readonly=True)
            log_info("Connessione al database stabilita con successo")
            return True
        except Exception as e:
            log_info(f"Errore di connessione al database: {str(e)}")
            return False

    def _get_optimized_query(self, view_name: str) -> str:
        """Genera una query ottimizzata"""
        return f'''
            SELECT *
            FROM "{view_name}"
            ORDER BY 1
        '''

    def fetch_view_data(self, view_name: str) -> pd.DataFrame:
        """Estrae i dati dalla vista in modo ottimizzato"""
        try:
            log_info(f"Inizio estrazione dati dalla vista: {view_name}")
            start_time = time.time()
            
            # Ottiene il conteggio totale
            count_query = f'SELECT COUNT(*) FROM "{view_name}"'
            total_rows = pd.read_sql_query(count_query, self.conn).iloc[0, 0]
            log_info(f"Totale righe da estrarre da {view_name}: {total_rows}")

            # Per viste piccole, estrazione diretta
            if total_rows < self.CHUNK_SIZE:
                query = self._get_optimized_query(view_name)
                df = pd.read_sql_query(query, self.conn)
                elapsed_time = time.time() - start_time
                log_info(f"Estratte {len(df)} righe da {view_name} in {elapsed_time:.2f} secondi")
                return df

            # Per viste grandi, usa COPY per massima performance
            buffer = io.StringIO()
            copy_sql = f'COPY (SELECT * FROM "{view_name}") TO STDOUT WITH CSV HEADER'
            
            with tqdm(total=total_rows, desc=f"Estrazione {view_name}", unit="righe", file=tqdm_logger) as pbar:
                cursor = self.conn.cursor()
                cursor.copy_expert(copy_sql, buffer)
                cursor.close()
                
                # Reset buffer e leggi con pandas
                buffer.seek(0)
                df = pd.read_csv(buffer)
                pbar.update(total_rows)

            elapsed_time = time.time() - start_time
            log_info(f"Estratte {len(df)} righe da {view_name} in {elapsed_time:.2f} secondi")
            return df

        except Exception as e:
            log_info(f"Errore nell'estrazione dalla vista {view_name}: {str(e)}")
            return pd.DataFrame()

    def close(self):
        if self.conn:
            self.conn.close()
            log_info("Connessione al database chiusa")

def export_to_excel(environment_config: Dict, output_path: str = None, environment_name: str = "UNKNOWN") -> str:
    """
    Export database views to Excel file
    
    Args:
        environment_config: Dictionary containing database configuration and views
        output_path: Optional path for the output file. If None, generates a default path
        environment_name: Name of the environment (e.g., SIT, UAT, PREPROD)
    
    Returns:
        str: Path to the generated Excel file
    """
    # Definisci il percorso base per le estrazioni
    base_path = Path(r"\\tsclient\V\Estrazioni")
    
    try:
        # Crea la directory se non esiste
        base_path.mkdir(parents=True, exist_ok=True)
        log_info(f"Directory di output verificata: {base_path}")
    except Exception as e:
        log_info(f"Impossibile creare la directory {base_path}. Uso il percorso locale. Errore: {str(e)}")
        base_path = Path.cwd()

    if output_path is None:
        # Crea sottodirectory per anno e mese
        current_date = datetime.now()
        year_month = current_date.strftime("%Y_%m")
        year_month_path = base_path / year_month
        
        try:
            year_month_path.mkdir(parents=True, exist_ok=True)
            log_info(f"Creata directory per anno/mese: {year_month_path}")
        except Exception as e:
            log_info(f"Impossibile creare la directory {year_month_path}. Uso il percorso base. Errore: {str(e)}")
            year_month_path = base_path

        timestamp = current_date.strftime("%Y%m%d_%H%M%S")
        output_path = str(year_month_path / f"export_{environment_name}_{timestamp}.xlsx")

    try:
        log_info(f"Inizializzazione export per ambiente {environment_name}")
        excel_writer = ExcelWriter()
        excel_writer.initialize(output_path)

        total_views = sum(len(db_config["views"]) for db_config in environment_config)
        progress_bar = tqdm(total=total_views, desc="Progresso totale", unit="vista", file=tqdm_logger)

        # Prima elabora tutte le tabelle piccole
        log_info("Fase 1: Elaborazione tabelle piccole")
        for db_config in environment_config:
            fetcher = DatabaseFetcher(db_config["config"])
            if not fetcher.connect():
                progress_bar.update(len(db_config["views"]))
                continue

            try:
                for view_name in db_config["views"]:
                    if view_name.lower() not in LARGE_TABLES:
                        start_time = time.time()
                        df = fetcher.fetch_view_data(view_name)
                        
                        if not df.empty:
                            excel_writer.write_dataframe(df, view_name)
                        
                        elapsed_time = time.time() - start_time
                        log_info(f"Elaborazione {view_name} completata in {elapsed_time:.2f} secondi")
                        
                        progress_bar.update(1)
                        progress_bar.set_description(f"Completata {view_name}")
                        
                        # Forza pulizia memoria
                        del df
                        gc.collect()
            finally:
                fetcher.close()

        # Poi elabora le tabelle grandi
        log_info("Fase 2: Elaborazione tabelle grandi")
        for db_config in environment_config:
            fetcher = DatabaseFetcher(db_config["config"])
            if not fetcher.connect():
                progress_bar.update(len(db_config["views"]))
                continue

            try:
                for view_name in db_config["views"]:
                    if view_name.lower() in LARGE_TABLES:
                        start_time = time.time()
                        df = fetcher.fetch_view_data(view_name)
                        
                        if not df.empty:
                            excel_writer.write_dataframe(df, view_name)
                        
                        elapsed_time = time.time() - start_time
                        log_info(f"Elaborazione {view_name} completata in {elapsed_time:.2f} secondi")
                        
                        progress_bar.update(1)
                        progress_bar.set_description(f"Completata {view_name}")
                        
                        # Forza pulizia memoria
                        del df
                        gc.collect()
            finally:
                fetcher.close()

        progress_bar.close()
        
        # Fase finale: unione dei file
        log_info("Fase 3: Unione dei file")
        excel_writer.finalize()
        log_info(f"Export completato. File salvato in: {output_path}")
        return output_path
        
    except PermissionError as e:
        log_info(f"Errore di permessi durante il salvataggio in {output_path}. Assicurarsi di avere i permessi necessari.")
        raise
    except Exception as e:
        log_info(f"Errore durante l'export: {str(e)}")
        raise
